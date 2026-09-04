# -*- coding: utf-8 -*-
import os
import json
from odoo import api, fields, models, _
# pyrefly: ignore [missing-import]
# pyrefly: ignore [missing-import]
from odoo.exceptions import ValidationError
import base64
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class WujiaFranchiseInspectionCategory(models.Model):
    _name = 'wujia.franchise.inspection.category'
    _description = 'Inspection Category / Group'
    _order = 'sequence, id'

    name = fields.Char(string='Category Name', required=True)
    code = fields.Char(string='Category Code')
    sequence = fields.Integer(string='Sequence', default=10)
    active = fields.Boolean(string='Active', default=True)
    is_severe = fields.Boolean(string='Severe Violation', default=False)

    _sql_constraints = [
        ('name_unique', 'UNIQUE(name)', 'Category Name must be unique!'),
    ]

    @api.model
    def _init_default_categories(self):
        """Khởi tạo các Danh mục tiêu chí mặc định nếu chưa tồn tại trong DB."""
        default_categories = [
            {'name': 'Gìn giữ hình ảnh ngoại quan cửa hàng / 店鋪外觀形象保持', 'sequence': 10},
            {'name': 'Yêu cầu giữ gìn các thiết bị / 各設備維護要求', 'sequence': 20},
            {'name': 'Yêu cầu tiêu chuẩn cơ bản / 基本規範要求', 'sequence': 30},
            {'name': 'Những hạng mục vi phạm nghiêm trọng (vi phạm 1 hạng mục bất kỳ sẽ bị trừ trực tiếp 6 điểm) / 任何一項嚴重違規,就直接扣六分', 'sequence': 40, 'is_severe': True},
        ]
        for cat_data in default_categories:
            existing = self.search([('name', '=', cat_data['name'])], limit=1)
            if not existing:
                self.create(cat_data)



class WujiaFranchiseInspectionTemplate(models.Model):
    _name = 'wujia.franchise.inspection.template'
    _description = 'Franchise Store Inspection Template'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'effective_date desc, id desc'

    name = fields.Char(string='Template Name', required=True, tracking=True)
    code = fields.Char(string='Template Code', tracking=True)
    version = fields.Char(string='Version', default='v1.0', tracking=True)
    state = fields.Selection([
        ('draft', 'Draft'),
        ('active', 'Active'),
        ('archived', 'Archived'),
    ], string='Status', default='draft', required=True, tracking=True)
    effective_date = fields.Date(string='Effective Date', default=fields.Date.context_today, tracking=True)
    
    checklist_max_score = fields.Float(string='Checklist Max Score', default=95.0, tracking=True)
    exam_max_score = fields.Float(string='Exam Max Score', default=5.0, tracking=True)
    total_max_score = fields.Float(
        string='Total Max Score',
        compute='_compute_total_max_score',
        store=True,
        tracking=True,
    )
    
    line_ids = fields.One2many(
        'wujia.franchise.inspection.template.line',
        'template_id',
        string='Inspection Criteria',
        copy=True,
    )

    @api.depends('checklist_max_score', 'exam_max_score')
    def _compute_total_max_score(self):
        for rec in self:
            rec.total_max_score = (rec.checklist_max_score or 0.0) + (rec.exam_max_score or 0.0)

    @api.ondelete(at_uninstall=False)
    def _unlink_except_draft(self):
        for rec in self:
            if rec.state != 'draft':
                raise ValidationError(_('Only Draft templates can be deleted!'))

    def write(self, vals):
        if not self.env.context.get('bypass_state_check'):
            for rec in self:
                if rec.state in ('active', 'archived'):
                    allowed_keys = {'state', 'version', 'effective_date', 'name'}
                    if not set(vals.keys()).issubset(allowed_keys):
                        raise ValidationError(_(
                            'Inspection template "%s" is in "%s" status and cannot be modified!\n'
                            'If you want to change criteria, please click "Create New Version".'
                        ) % (rec.name, rec.state))
        return super().write(vals)

    def action_activate(self):
        for rec in self:
            if rec.code:
                old_actives = self.search([
                    ('id', '!=', rec.id),
                    ('code', '=', rec.code),
                    ('state', '=', 'active'),
                ])
                old_actives.write({'state': 'archived'})
            rec.write({'state': 'active'})

    def action_archive(self):
        self.write({'state': 'archived'})

    def action_reset_draft(self):
        self.write({'state': 'draft'})

    def _increment_version(self, ver_str):
        if not ver_str:
            return 'v1.1'
        clean_ver = ver_str.lower().lstrip('v')
        try:
            parts = clean_ver.split('.')
            if len(parts) >= 2:
                parts[-1] = str(int(parts[-1]) + 1)
                return f"v{'.'.join(parts)}"
            elif len(parts) == 1:
                return f"v{int(parts[0]) + 1}.0"
        except Exception:
            pass
        return f"{ver_str}-v2"

    def action_create_new_version(self):
        self.ensure_one()
        old_version = self.version or 'v1.0'
        new_version = self._increment_version(old_version)

        # 1. Tạo bản sao lưu (Clone) cho phiên bản cũ ở trạng thái draft, sau đó lưu trữ (archived)
        archived_copy = self.copy({
            'name': f"{self.name} ({old_version})",
            'version': old_version,
            'state': 'draft',
            'effective_date': self.effective_date,
        })
        if not archived_copy.line_ids and self.line_ids:
            new_lines = []
            for line in self.line_ids:
                new_lines.append({
                    'template_id': archived_copy.id,
                    'sequence': line.sequence,
                    'criterion_code': line.criterion_code,
                    'category_id': line.category_id.id if line.category_id else False,
                    'display_type': line.display_type,
                    'content': line.content,
                    'criterion_type': line.criterion_type,
                    'deduction_score': line.deduction_score,
                    'require_note_if_fail': line.require_note_if_fail,
                    'require_evidence_if_fail': line.require_evidence_if_fail,
                    'active': line.active,
                })
            if new_lines:
                self.env['wujia.franchise.inspection.template.line'].create(new_lines)

        # Chuyển bản clone cũ sang trạng thái archived
        archived_copy.write({'state': 'archived'})

        # 2. Cập nhật bản ghi hiện tại lên Version mới (GIỮ NGUYÊN ID VÀ ID CÁC DÒNG TIÊU CHÍ HIỆN TẠI)
        self.with_context(bypass_state_check=True).write({
            'version': new_version,
            'state': 'draft',
            'effective_date': fields.Date.context_today(self),
        })

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Nâng cấp phiên bản thành công / Version Updated'),
                'message': _('Đã lưu trữ phiên bản cũ (%s) và nâng bản hiện tại lên %s (Giữ nguyên toàn bộ ID tiêu chí).') % (old_version, new_version),
                'type': 'success',
                'sticky': False,
            }
        }

    def action_open_import_wizard(self):
        """Mở popup wizard nhập tiêu chí từ Excel."""
        self.ensure_one()
        if self.state in ('active', 'archived'):
            raise ValidationError(_(
                'Không thể nhập tiêu chí cho Mẫu ở trạng thái "%s"!\n'
                'Vui lòng nhấn "Tạo phiên bản mới" hoặc "Chuyển về Dự thảo" để chỉnh sửa tiêu chí.'
            ) % (self.state))
        return {
            'name': _('Nhập tiêu chí từ Excel / Import Criteria'),
            'type': 'ir.actions.act_window',
            'res_model': 'wujia.franchise.inspection.template.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_template_id': self.id,
            },
        }

    def _generate_excel_workbook(self, is_sample=False):
        """Sinh workbook openpyxl cho template (dùng để tải mẫu hoặc xuất tiêu chí hiện tại)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tieu_Chi_Khao_Sat"

        # Styles
        header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        section_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        section_font = Font(name="Calibri", size=11, bold=True, color="1B5E20")
        line_font = Font(name="Calibri", size=10, color="000000")
        thin_border = Border(
            left=Side(style='thin', color='D0D0D0'),
            right=Side(style='thin', color='D0D0D0'),
            top=Side(style='thin', color='D0D0D0'),
            bottom=Side(style='thin', color='D0D0D0')
        )
        align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

        headers = [
            "STT",
            "Mã tiêu chí",
            "Loại dòng (Section/Line)",
            "Danh mục / Nhóm",
            "Nội dung tiêu chí",
            "Phân loại (Thông thường/Điểm liệt)",
            "Deduction Score",
            "Bắt buộc ghi chú nếu vi phạm (Có/Không)",
            "Bắt buộc ảnh nếu vi phạm (Có/Không)"
        ]

        # Write Header
        ws.row_dimensions[1].height = 28
        for col_num, header_title in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_title)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = thin_border

        if is_sample:
            # Sample Data
            sample_rows = [
                (10, "1.", "Section", "Gìn giữ hình ảnh ngoại quan cửa hàng / 店鋪外觀形象保持", "", "", 0.0, "Không", "Không"),
                (11, "1.01", "Line", "Gìn giữ hình ảnh ngoại quan cửa hàng / 店鋪外觀形象保持", "Duy trì trang thiết bị ngoại quan: Bảng hiệu, đèn chiếu sáng, mái hiên và băng rôn \"Mua 5 tặng 1\" phải giữ gìn sạch sẽ, không bị phai màu hay hư hỏng.\n門外設施維護:招牌、照明、雨棚及「買五送一」活動布條需保持完好,不得有褪色,破損或明顯髒污。", "Điểm liệt", 2.0, "Không", "Có"),
                (12, "1.02", "Line", "Gìn giữ hình ảnh ngoại quan cửa hàng / 店鋪外觀形象保持", "Phía trước tiệm và phía dưới dốc dắt xe máy sạch rác và lá cây, mặt đất giữ sạch sẽ, không có bụi bẩn hay vết ố.\n店前及機車坡道下方地面整潔無垃圾、落葉,地面保持乾淨無髒污。", "Thông thường", 1.0, "Không", "Không"),
                (13, "1.03", "Line", "Gìn giữ hình ảnh ngoại quan cửa hàng / 店鋪外觀形象保持", "Xe đậu ngay ngắn; đảm bảo lối ra vào tiệm và khu vực quanh xe trà luôn thông thoáng, không có vật cản gây trở ngại cho khách mua hàng.\n車輛停放整齊;確保店鋪進出口及茶攤車周邊動線通暢,嚴禁擺放任何阻礙顧客購餐之障礙物。", "Thông thường", 1.0, "Không", "Không"),
                (20, "2.", "Section", "Yêu cầu giữ gìn các thiết bị / 各設備維護要求", "", "", 0.0, "Không", "Không"),
                (21, "2.01", "Line", "Yêu cầu giữ gìn các thiết bị / 各設備維護要求", "Bình đựng trà, máy đo định lượng đường, máy làm đá và các thiết bị pha chế được vệ sinh sạch sẽ, hoạt động bình thường.\n茶桶、果糖機、製冰機等吧台設備保持清潔且運作正常。", "Thông thường", 1.0, "Không", "Không"),
                (30, "3.", "Section", "Những hạng mục vi phạm nghiêm trọng (trừ 6 điểm) / 嚴重違規項", "", "", 0.0, "Không", "Không"),
                (31, "3.01", "Line", "Những hạng mục vi phạm nghiêm trọng (trừ 6 điểm) / 嚴重違規項", "Sử dụng nguyên liệu không rõ nguồn gốc hoặc hết hạn sử dụng.\n使用來源不明或過期之原物料。", "Điểm liệt", 6.0, "Có", "Có"),
            ]
            for row_idx, r_data in enumerate(sample_rows, 2):
                is_sec = (r_data[2].lower() == "section")
                ws.row_dimensions[row_idx].height = 24 if is_sec else 45
                for col_idx, val in enumerate(r_data, 1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.font = section_font if is_sec else line_font
                    if is_sec:
                        c.fill = section_fill
                    c.border = thin_border
                    c.alignment = align_center if col_idx in (1, 2, 3, 6, 7, 8, 9) else align_left
        else:
            # Export existing template lines
            for row_idx, line in enumerate(self.line_ids, 2):
                is_sec = (line.display_type == 'section')
                ws.row_dimensions[row_idx].height = 24 if is_sec else 45
                cat_name = line.category_id.name if line.category_id else ''
                row_vals = [
                    line.sequence or 10,
                    line.criterion_code or '',
                    'Section' if is_sec else 'Line',
                    cat_name,
                    line.content or '' if not is_sec else '',
                    'Điểm liệt' if line.criterion_type == 'critical' else 'Thông thường',
                    line.deduction_score or 0.0,
                    'Có' if line.require_note_if_fail else 'Không',
                    'Có' if line.require_evidence_if_fail else 'Không',
                ]
                for col_idx, val in enumerate(row_vals, 1):
                    c = ws.cell(row=row_idx, column=col_idx, value=val)
                    c.font = section_font if is_sec else line_font
                    if is_sec:
                        c.fill = section_fill
                    c.border = thin_border
                    c.alignment = align_center if col_idx in (1, 2, 3, 6, 7, 8, 9) else align_left

        # Adjust column widths
        col_widths = {
            1: 8,    # STT
            2: 14,   # Mã tiêu chí
            3: 16,   # Loại dòng
            4: 35,   # Danh mục
            5: 65,   # Nội dung
            6: 18,   # Phân loại
            7: 12,   # Điểm trừ
            8: 18,   # Bắt buộc ghi chú
            9: 18,   # Bắt buộc ảnh
        }
        for col_idx, width in col_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def action_download_excel_sample(self):
        """Tải file Excel mẫu tiêu chí chuẩn."""
        file_bytes = self._generate_excel_workbook(is_sample=True)
        file_name = "Mau_Tieu_Chi_Khao_Sat_Wujia.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'type': 'binary',
            'datas': base64.b64encode(file_bytes),
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    def action_export_criteria_excel(self):
        """Xuất toàn bộ tiêu chí của mẫu hiện tại ra file Excel."""
        self.ensure_one()
        file_bytes = self._generate_excel_workbook(is_sample=False)
        code_str = self.code or self.name or 'Template'
        code_clean = "".join(c for c in code_str if c.isalnum() or c in ('_', '-')).strip()
        file_name = f"Tieu_Chi_{code_clean}_{self.version or 'v1'}.xlsx"
        attachment = self.env['ir.attachment'].create({
            'name': file_name,
            'type': 'binary',
            'datas': base64.b64encode(file_bytes),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        })
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{attachment.id}?download=true',
            'target': 'self',
        }

    @api.model
    def _init_demo_template(self):
        """Khởi tạo dữ liệu mẫu khảo sát MM01 (Khảo sát cửa hàng nhượng quyền) nếu chưa tồn tại trong DB."""
        existing = self.search([('code', '=', 'MM01')], limit=1)
        if existing:
            return

        Category = self.env['wujia.franchise.inspection.category']

        template = self.create({
            'name': 'Khảo sát cửa hàng nhượng quyền',
            'code': 'MM01',
            'state': 'draft',
            'version': 'v1.0',
            'checklist_max_score': 95.0,
            'exam_max_score': 5.0,
            'total_max_score': 100.0,
            'effective_date': fields.Date.today(),
        })

        json_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'template_mm01_lines.json')
        if os.path.exists(json_path):
            with open(json_path, 'r', encoding='utf-8') as f:
                lines_data = json.load(f)

            category_cache = {}
            lines_to_create = []

            for item in lines_data:
                cat_name = item.get('category_name')
                cat_id = False
                if cat_name:
                    if cat_name not in category_cache:
                        cat_rec = Category.search([('name', '=', cat_name)], limit=1)
                        if not cat_rec:
                            cat_rec = Category.create({'name': cat_name, 'sequence': 10})
                        category_cache[cat_name] = cat_rec.id
                    cat_id = category_cache[cat_name]

                lines_to_create.append({
                    'template_id': template.id,
                    'sequence': item.get('sequence', 10),
                    'display_type': item.get('display_type') or 'line',
                    'criterion_code': item.get('criterion_code'),
                    'category_id': cat_id,
                    'content': item.get('content'),
                    'criterion_type': item.get('criterion_type') or 'normal',
                    'deduction_score': item.get('deduction_score', 1.0),
                    'require_note_if_fail': item.get('require_note_if_fail', False),
                    'require_evidence_if_fail': item.get('require_evidence_if_fail', False),
                })

            if lines_to_create:
                self.env['wujia.franchise.inspection.template.line'].create(lines_to_create)

        template.write({'state': 'active'})


class WujiaFranchiseInspectionTemplateLine(models.Model):
    _name = 'wujia.franchise.inspection.template.line'
    _description = 'Inspection Template Line'
    _order = 'sequence, id'

    template_id = fields.Many2one(
        'wujia.franchise.inspection.template',
        string='Inspection Template',
        required=True,
        ondelete='cascade',
    )
    sequence = fields.Integer(string='Sequence', default=10)
    display_type = fields.Selection([
        ('line', 'Criterion Line'),
        ('section', 'Section Header'),
    ], string='Display Type', default='line', required=True)
    criterion_code = fields.Char(string='Criterion Code')
    category_id = fields.Many2one(
        'wujia.franchise.inspection.category',
        string='Category',
    )
    category = fields.Char(string='Category (Legacy)', related='category_id.name', readonly=True, store=True)
    content = fields.Text(string='Criterion Content')
    criterion_type = fields.Selection([
        ('normal', 'Normal'),
        ('critical', 'Critical / Knockout'),
    ], string='Criterion Type', default='normal', required=True)
    deduction_score = fields.Float(string='Deduction Score', default=1.0)
    require_note_if_fail = fields.Boolean(string='Require Note If Fail', default=False)
    require_evidence_if_fail = fields.Boolean(string='Require Evidence If Fail', default=False)
    active = fields.Boolean(string='Active', default=True)
    is_severe = fields.Boolean(string='Severe Violation', default=False)

    @api.onchange('category_id', 'display_type')
    def _onchange_category_id_section(self):
        """Khi chọn loại là Section và chọn Danh mục, tự điền Nội dung = Tên danh mục."""
        if self.display_type == 'section' and self.category_id:
            self.content = self.category_id.name

    @api.depends('criterion_code', 'content', 'display_type', 'category_id')
    def _compute_display_name(self):
        for rec in self:
            code_raw = (rec.criterion_code or '').strip()
            sub_code = code_raw.split('.', 1)[1].strip() if '.' in code_raw else ''
            if rec.display_type == 'section':
                cat_name = rec.category_id.name if rec.category_id else rec.content
                if sub_code:
                    rec.display_name = f"[{sub_code}] {cat_name or _('No Category Selected')}"
                else:
                    rec.display_name = cat_name or _('No Category Selected')
            else:
                content = (rec.content or '').strip()
                if sub_code:
                    rec.display_name = f"[{sub_code}] {content}"
                else:
                    rec.display_name = content or _("Unnamed Criterion")

    @api.ondelete(at_uninstall=False)
    def _unlink_except_draft_template(self):
        for line in self:
            if line.template_id.state in ('active', 'archived'):
                raise ValidationError(_('Cannot delete criteria of an Active or Archived template!'))

    def write(self, vals):
        for line in self:
            if line.template_id.state in ('active', 'archived'):
                raise ValidationError(_('Cannot modify criteria of an Active or Archived template!'))
        return super().write(vals)

    @api.model_create_multi
    def create(self, vals_list):
        lines = super().create(vals_list)
        for line in lines:
            if line.template_id.state in ('active', 'archived'):
                raise ValidationError(_('Cannot add new criteria to an Active or Archived template!'))
        return lines


class WujiaFranchiseInspectionTemplateImportWizard(models.TransientModel):
    _name = 'wujia.franchise.inspection.template.import.wizard'
    _description = 'Import Inspection Criteria Wizard'

    template_id = fields.Many2one(
        'wujia.franchise.inspection.template',
        string='Mẫu khảo sát',
        required=True,
        ondelete='cascade',
    )
    file_data = fields.Binary(string='File Excel (.xlsx)', required=True)
    file_name = fields.Char(string='Tên file', default='criteria_import.xlsx')
    import_mode = fields.Selection([
        ('replace', 'Xóa cũ và thay thế toàn bộ (Replace all existing criteria)'),
        ('append', 'Thêm tiếp vào danh sách hiện tại (Append to current list)'),
    ], string='Chế độ nhập', default='replace', required=True)
    auto_create_category = fields.Boolean(
        string='Tự động tạo Danh mục nếu chưa có',
        default=True,
        help='Nếu tên Danh mục trong file Excel chưa tồn tại trong hệ thống, hệ thống sẽ tự động tạo mới.'
    )

    def action_download_sample_excel(self):
        """Tải file mẫu từ wizard."""
        return self.template_id.action_download_excel_sample()

    def action_import_excel(self):
        self.ensure_one()
        if not self.file_data:
            raise ValidationError(_('Vui lòng chọn file Excel để nhập!'))

        if self.file_name and not self.file_name.lower().endswith(('.xlsx', '.xlsm')):
            raise ValidationError(_('Chỉ hỗ trợ file Excel định dạng .xlsx hoặc .xlsm!'))

        if self.template_id.state in ('active', 'archived'):
            raise ValidationError(_(
                'Mẫu khảo sát "%s" đang ở trạng thái "%s" không thể chỉnh sửa tiêu chí!\n'
                'Vui lòng bấm "Tạo phiên bản mới" hoặc "Chuyển về Dự thảo".'
            ) % (self.template_id.name, self.template_id.state))

        try:
            file_bytes = base64.b64decode(self.file_data)
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        except Exception as e:
            raise ValidationError(_('Không thể đọc file Excel. Lỗi: %s') % str(e))

        ws = wb.active
        if not ws:
            raise ValidationError(_('File Excel không có trang tính (sheet) nào!'))

        Category = self.env['wujia.franchise.inspection.category']
        category_cache = {}
        for cat in Category.search([]):
            category_cache[cat.name.strip().lower()] = cat.id

        parsed_rows = []
        row_count = 0
        section_count = 0
        line_count = 0
        current_cat_id = False
        default_seq = 10
        if self.import_mode == 'append' and self.template_id.line_ids:
            max_seq = max(self.template_id.line_ids.mapped('sequence') or [0])
            default_seq = max_seq + 10

        # Xác định dòng bắt đầu (bỏ qua header nếu có)
        start_row = 1
        first_row_vals = [str(ws.cell(row=1, column=c).value or '').strip().lower() for c in range(1, 10)]
        header_keywords = ['stt', 'mã', 'ma', 'code', 'loại', 'loai', 'danh mục', 'danh muc', 'nội dung', 'noi dung', 'content', 'tiêu chí', 'tieu chi']
        if any(kw in " ".join(first_row_vals) for kw in header_keywords):
            start_row = 2

        for row_idx in range(start_row, ws.max_row + 1):
            seq_val = ws.cell(row=row_idx, column=1).value
            code_val = ws.cell(row=row_idx, column=2).value
            type_val = ws.cell(row=row_idx, column=3).value
            cat_val = ws.cell(row=row_idx, column=4).value
            content_val = ws.cell(row=row_idx, column=5).value
            crit_type_val = ws.cell(row=row_idx, column=6).value
            deduct_val = ws.cell(row=row_idx, column=7).value
            req_note_val = ws.cell(row=row_idx, column=8).value
            req_img_val = ws.cell(row=row_idx, column=9).value

            # Xử lý chuỗi
            code_str = str(code_val).strip() if code_val is not None else ''
            if code_str.endswith('.0') and not '.' in code_str[:-2]:
                # Xử lý trường hợp int đọc thành float
                try:
                    code_str = str(int(float(code_str)))
                except Exception:
                    pass

            type_str = str(type_val).strip().lower() if type_val is not None else ''
            cat_str = str(cat_val).strip() if cat_val is not None else ''
            content_str = str(content_val).strip() if content_val is not None else ''

            # Bỏ qua dòng trống
            if not code_str and not cat_str and not content_str:
                continue

            # Xử lý Category
            if cat_str:
                cat_key = cat_str.lower()
                if cat_key not in category_cache:
                    if self.auto_create_category:
                        new_cat = Category.create({
                            'name': cat_str,
                            'sequence': len(category_cache) * 10 + 10,
                        })
                        category_cache[cat_key] = new_cat.id
                        current_cat_id = new_cat.id
                    else:
                        current_cat_id = False
                else:
                    current_cat_id = category_cache[cat_key]

            # Xác định display_type (section vs line)
            is_section = False
            if type_str in ('section', 'phần', 'phan', 'mục', 'muc', 'hạng mục', 'hang muc', 'nhóm', 'nhom'):
                is_section = True
            elif (code_str.endswith('.') or '.' not in code_str) and not content_str:
                is_section = True
            elif content_str and cat_str and content_str.lower() == cat_str.lower():
                is_section = True

            display_type = 'section' if is_section else 'line'

            # Xử lý Sequence
            seq = default_seq
            if seq_val is not None:
                try:
                    seq = int(float(str(seq_val).strip()))
                except Exception:
                    seq = default_seq
            default_seq = seq + 1

            # Xử lý Criterion Type
            crit_type_str = str(crit_type_val).strip().lower() if crit_type_val is not None else ''
            if any(k in crit_type_str for k in ('critical', 'liệt', 'liet', 'nghiêm trọng', 'nghiem trong', 'knockout')):
                criterion_type = 'critical'
            else:
                criterion_type = 'normal'

            # Xử lý Deduction Score
            deduct_score = 1.0
            if deduct_val is not None:
                try:
                    deduct_score = float(str(deduct_val).strip())
                except Exception:
                    deduct_score = 0.0 if is_section else 1.0
            else:
                deduct_score = 0.0 if is_section else 1.0

            # Xử lý Boolean (require note / evidence)
            def parse_bool(v):
                if v is None:
                    return False
                v_str = str(v).strip().lower()
                return v_str in ('1', 'true', 'yes', 'có', 'co', 'x', 'v', 'y')

            require_note = parse_bool(req_note_val)
            require_img = parse_bool(req_img_val)

            parsed_rows.append({
                'sequence': seq,
                'criterion_code': code_str or False,
                'category_id': current_cat_id,
                'display_type': display_type,
                'content': content_str or (cat_str if is_section else False),
                'criterion_type': criterion_type,
                'deduction_score': deduct_score,
                'require_note_if_fail': require_note,
                'require_evidence_if_fail': require_img,
                'active': True,
            })

            if is_section:
                section_count += 1
            else:
                line_count += 1
            row_count += 1

        if not parsed_rows:
            raise ValidationError(_('Không tìm thấy dòng dữ liệu tiêu chí hợp lệ nào trong file Excel!'))

        # Lập chỉ mục các dòng tiêu chí đang có trong Template hiện tại
        existing_lines = self.template_id.line_ids
        existing_by_code = {}
        existing_sections = {}

        for el in existing_lines:
            if el.display_type == 'section':
                sec_key = (el.content or (el.category_id.name if el.category_id else '') or '').strip().lower()
                if sec_key and sec_key not in existing_sections:
                    existing_sections[sec_key] = el
            else:
                code_key = (el.criterion_code or '').strip().lower()
                if code_key and code_key not in existing_by_code:
                    existing_by_code[code_key] = el

        retained_line_ids = set()
        created_count = 0
        updated_count = 0

        # Tiến hành Upsert: CHỈ CHECK THEO MÃ TIÊU CHÍ (criterion_code)
        for r_vals in parsed_rows:
            matched_line = False
            is_sec = (r_vals.get('display_type') == 'section')

            if is_sec:
                sec_key = (r_vals.get('content') or '').strip().lower()
                if sec_key in existing_sections:
                    matched_line = existing_sections[sec_key]
            else:
                # Chỉ kiểm tra theo mã tiêu chí (criterion_code)
                code_key = (r_vals.get('criterion_code') or '').strip().lower()
                if code_key and code_key in existing_by_code:
                    matched_line = existing_by_code[code_key]

            if matched_line and matched_line.id not in retained_line_ids:
                # 1. Đúng mã tiêu chí -> CẬP NHẬT VÀ GIỮ NGUYÊN ID CŨ
                matched_line.write(r_vals)
                retained_line_ids.add(matched_line.id)
                updated_count += 1
            else:
                # 2. Sai mã / Mã mới -> TẠO BẢN GHI MỚI VỚI ID MỚI
                r_vals['template_id'] = self.template_id.id
                new_l = self.env['wujia.franchise.inspection.template.line'].create(r_vals)
                retained_line_ids.add(new_l.id)
                created_count += 1

        # Nếu chế độ là 'replace' (thay thế), xóa các dòng cũ không còn xuất hiện trong file Excel
        if self.import_mode == 'replace':
            redundant_lines = existing_lines.filtered(lambda l: l.id not in retained_line_ids)
            if redundant_lines:
                redundant_lines.unlink()

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Nhập Excel Thành Công / Success'),
                'message': _(
                    'Đã xử lý %d dòng: Cập nhật & Giữ nguyên ID (%d dòng trùng mã), Tạo mới (%d dòng mới).'
                ) % (row_count, updated_count, created_count),
                'type': 'success',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }

