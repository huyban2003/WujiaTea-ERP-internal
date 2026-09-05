#!/usr/bin/env python3
"""Dump các dòng liên quan từ BA master plan xlsm → text để đối chiếu model/field/feature.

Nguồn: docs/Wujia_Internal ERP Master Plan.xlsm. Sheet quan trọng khi làm controller:
  '1. Model Field'   → model/field backend (BA đặt tên lý tưởng hoá, vd wujia.announcement)
  '2. FE - Portal'   → hành vi frontend / sitemap
  '3. Controller'    → bảng CT-0xx (chức năng kết nối portal↔backend)
  'FEATURE CHECKLIST'→ POR-0xx (feature + trạng thái)

Dev-only — không thuộc module Odoo, không chạy trên server (thư mục này gitignored).

Dùng:
    python3 read_xlsm.py --sheets                         # liệt kê sheet
    python3 read_xlsm.py "1. Model Field" announcement priority
    python3 read_xlsm.py "3. Controller" "thông báo"
"""
import argparse
import os
import sys

try:
    import openpyxl
except ImportError:
    sys.exit('!! pip install openpyxl (env odoo đã có sẵn).')

DEFAULT_XLSM = os.path.join(
    os.path.dirname(__file__), '..', '..', 'docs',
    'Wujia_Internal ERP Master Plan.xlsm')


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('sheet', nargs='?', help='tên sheet')
    ap.add_argument('keywords', nargs='*', help='lọc dòng chứa BẤT KỲ keyword (case-insensitive)')
    ap.add_argument('--file', default=DEFAULT_XLSM, help='đường dẫn xlsm')
    ap.add_argument('--sheets', action='store_true', help='chỉ liệt kê tên sheet rồi thoát')
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    if args.sheets or not args.sheet:
        print('SHEETS:')
        for i, s in enumerate(wb.sheetnames):
            print(f'  [{i}] {s!r}')
        return
    if args.sheet not in wb.sheetnames:
        sys.exit(f'!! Sheet {args.sheet!r} không có. Chạy --sheets để xem danh sách.')

    ws = wb[args.sheet]
    kws = [k.lower() for k in args.keywords]
    rows = list(ws.iter_rows(values_only=True))
    kept = 0
    for ri, r in enumerate(rows):
        cells = [str(c).strip() for c in r if c not in (None, '')]
        if not cells:
            continue
        joined = ' '.join(cells).lower()
        if not kws or any(k in joined for k in kws):
            print(f'[r{ri}] ' + ' ┃ '.join(cells))
            kept += 1
    print(f'--- {kept} dòng khớp trong {args.sheet!r} ---', file=sys.stderr)


if __name__ == '__main__':
    main()
